"""
Excel üretim servisi (FULL Servis, sunucu tarafı) — ping ve wifi log dosyalarından
GRK ile BİREBİR aynı yapıda Excel raporları üretir.

GRK karşılıkları (port edildi):
  • Ping  → app/utils/wifi/analiz/pinganaliz/pingLogParser.py + merge_parser.py
            (BİLGİSAYAR BAŞINA tek ÖZET Excel: modem + internet, IPv4/IPv6 tüm ping
             logları tek dosyada birleşir — Statistics + Graphs + Data sayfaları,
             anomali tespiti, her hedef için ping süre grafiği)
  • Wifi  → app/utils/wifi/analiz/wifianaliz/parse.py + chart.py
            (Veriler + Özet + Grafikler sayfaları, 10 alt grafik). Bilgisayar başına
            tek wifi logu olduğu için wifi'de birleştirme yok, log→Excel birebir.

TEK FARK (zorunlu): GRK ping parser'ı yalnızca Windows ping formatını
("Reply from X: ... time=12ms") tanır. FULL Servis Linux sunucu ve iki Mac'te de
ping attığı için parser CROSS-PLATFORM yapıldı — Windows + Unix (Linux/macOS)
çıktısını da okur ("64 bytes from X: ... time=12.3 ms").

Ağır bağımlılıklar (pandas/numpy/matplotlib/xlsxwriter/openpyxl) yalnızca BU dosyada
ve yalnızca sunucuda gerekir. Import'lar fonksiyon içinde (lazy) yapılır; kurulu
değillerse Excel üretilmez ama sistem ÇALIŞMAYA DEVAM EDER (sadece log .txt gider).

Java notu: Buradaki her fonksiyon "girdi log dosyası → çıktı .xlsx yolu" üreten saf
bir dönüştürücüdür (bir Service metodu gibi). DataFrame = tablo (ResultSet benzeri),
Figure = bellek-içi grafik nesnesi (sonra PNG'ye basılıp Excel'e gömülür).
"""
from __future__ import annotations

import os
import re
from collections import Counter
from datetime import datetime
from io import BytesIO


# ─────────────────────────────────────────────────────────────────────────────
# PING — GRK pingLogParser.py portu (cross-platform parse + aynı Excel yapısı)
# ─────────────────────────────────────────────────────────────────────────────

def _parse_ping_log(file_path):
    """Bir ping log dosyasını ayrıştırıp (DataFrame, target_ip) döner.

    Windows  : 'Reply from 8.8.8.8: bytes=32 time=12ms TTL=64' / 'Request timed out'
    Unix     : '64 bytes from 8.8.8.8: icmp_seq=1 ttl=53 time=12.3 ms'
    (Unix'te kayıp paket genelde satır basmaz; eldeki yanıtlardan istatistik çıkar.)
    """
    import pandas as pd

    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        lines = f.readlines()

    # Hedef IP: Windows "Pinging X", Unix "PING X (", ya da FULL Servis başlığı "Hedef: X"
    target_ip = None
    for line in lines:
        m = (re.search(r"Pinging\s+([0-9a-fA-F:.]+)", line)
             or re.search(r"PING\s+([0-9a-fA-F:.]+)", line)
             or re.search(r"Hedef:\s*([0-9a-fA-F:.]+)", line))
        if m:
            target_ip = m.group(1)
            break

    # Başarılı yanıttan zaman çek: "time=12ms", "time<1ms", "time=12.3 ms", TR "süre=12 ms"
    time_re = re.compile(r"(?:time|s[üu]re|zaman)\s*[=<]\s*([\d.,]+)\s*ms", re.IGNORECASE)
    reply_hint = ("reply from", "bytes from", "icmp_seq", "ttl=", "ttl ")

    ping_data = []
    seq = 0
    for line in lines:
        low = line.lower()

        # Zaman aşımı / ulaşılamaz (Windows ve bazı Unix sürümleri)
        if ("request timed out" in low or "destination host unreachable" in low
                or "host unreachable" in low):
            ping_data.append({"sequence": seq, "time_ms": None,
                              "status": "timeout", "raw_data": line.strip()})
            seq += 1
            continue

        # Başarılı yanıt (yanıt ipucu + zaman değeri birlikte olmalı)
        m = time_re.search(line)
        if m and any(h in low for h in reply_hint):
            try:
                t = float(m.group(1).replace(",", "."))
            except ValueError:
                continue
            ping_data.append({"sequence": seq, "time_ms": t,
                              "status": "success", "raw_data": line.strip()})
            seq += 1

    df = pd.DataFrame(ping_data)
    if not df.empty:
        df["target_ip"] = target_ip
    return df, target_ip


def _ping_statistics(df):
    """GRK calculate_statistics ile aynı istatistik sözlüğü."""
    if df is None or df.empty:
        return {}
    success_df = df[df["status"] == "success"]
    total = len(df)
    success = len(success_df)
    stats = {
        "target_ip": df["target_ip"].iloc[0] if "target_ip" in df.columns else "Unknown",
        "total_pings": total,
        "successful_pings": success,
        "failed_pings": total - success,
        "success_rate": (success / total * 100) if total else 0,
    }
    if not success_df.empty:
        stats.update({
            "min_time": success_df["time_ms"].min(),
            "max_time": success_df["time_ms"].max(),
            "avg_time": success_df["time_ms"].mean(),
            "median_time": success_df["time_ms"].median(),
            "std_dev": success_df["time_ms"].std(),
        })
    stats["packet_loss_percentage"] = 100 - stats["success_rate"]
    return stats


def _ping_identify_anomalies(df, z_threshold=3.0):
    """GRK identify_anomalies portu — z-skor ile anormal gecikme + tüm timeout'lar."""
    if df is None or df.empty:
        return df
    df = df.copy()
    df["is_anomaly"] = False
    success_mask = df["status"] == "success"
    if success_mask.sum() >= 2:
        times = df.loc[success_mask, "time_ms"]
        mean_t, std_t = times.mean(), times.std()
        if std_t and std_t > 0:
            z = abs((times - mean_t) / std_t)
            df.loc[success_mask, "is_anomaly"] = z > z_threshold
    df.loc[df["status"] == "timeout", "is_anomaly"] = True
    return df


def _ping_time_graph(df):
    """GRK create_ping_time_graph portu — anomalileri vurgulayan ping süre grafiği."""
    from matplotlib.figure import Figure

    fig = Figure(figsize=(15, 9))
    if df is None or df.empty:
        return fig
    ax = fig.add_subplot(111)

    success_df = df[df["status"] == "success"]
    timeout_df = df[df["status"] == "timeout"]

    if not success_df.empty:
        normal_df = success_df[~success_df["is_anomaly"]]
        if not normal_df.empty:
            ax.plot(normal_df["sequence"], normal_df["time_ms"], "-",
                    color="blue", label="Normal", alpha=0.7, linewidth=0.8)
        anomaly_df = success_df[success_df["is_anomaly"]]
        if not anomaly_df.empty:
            ax.plot(anomaly_df["sequence"], anomaly_df["time_ms"], "o",
                    color="red", label="Anomali", markersize=0.8)

    if not timeout_df.empty:
        y_max = (success_df["time_ms"].max() * 1.1) if not success_df.empty else 100
        ax.plot(timeout_df["sequence"], [y_max] * len(timeout_df), "x",
                color="red", markersize=5.0, label="Zaman Aşımı")

    ax.set_xlabel("Sıra")
    ax.set_ylabel("Yanıt Süresi (ms)")
    ax.set_title("Ping Yanıt Süreleri")
    ax.grid(True, linestyle="--", alpha=0.7)
    ax.legend()
    if "target_ip" in df.columns and df["target_ip"].iloc[0]:
        fig.text(0.02, 0.02, f"Hedef IP: {df['target_ip'].iloc[0]}", fontsize=10)
    return fig


def _ip_version(ip: str) -> str:
    """IP adresinde ':' varsa 'IPv6', aksi halde 'IPv4' (GRK _detect_ip_version portu)."""
    return "IPv6" if (ip and ":" in ip) else "IPv4"


# Statistics sayfasındaki sütunlar — GRK merge_parser.stats_keys ile birebir,
# başına iki FULL Servis alanı: hangi bilgisayar + hangi ham log dosyası.
_PING_STATS_KEYS = [
    "source_computer", "source_file", "target_ip", "ip_version",
    "total_pings", "successful_pings", "failed_pings",
    "success_rate", "packet_loss_percentage",
    "avg_time", "min_time", "max_time", "std_dev",
    "test_start_time", "test_end_time",
]

# Data sayfası satır tavanı — Telegram 50 MB sınırını aşmamak için (GRK ile aynı)
_PING_MAX_DATA_ROWS = 200_000


def ping_summary_excel(txt_paths: list[str], node_name: str,
                       brand=None, model=None, firmware=None,
                       out_dir: str | None = None,
                       test_start_time=None) -> str | None:
    """TEK BİR BİLGİSAYARIN tüm ping loglarını (modem + internet, IPv4/IPv6) birleştirip
    tek bir ÖZET Excel'i üretir — GRK merge_parser.merge_ping_logs portu.

    Sayfalar (GRK ile birebir):
      • Statistics : her ham log için bir satır (hedef IP, kayıp %, min/max/avg/std…)
      • Graphs     : her hedef için ping süre grafiği (anomaliler işaretli), alt alta
      • Data       : tüm logların ham satırları (source_computer/source_file sütunlu,
                     çok uzun testlerde eşit aralıklı downsample)

    İsimlendirme GRK ile birebir, TEK farkı araya giren bilgisayar (client) adı:
        GRK  : grk_ping_ozet_<marka>_<model>_<fw>_<zaman>.xlsx
        BURADA: FULL_Service_ping_ozet_<BILGISAYAR>_<marka>_<model>_<fw>_<zaman>.xlsx

    Üretilen Excel'in yolunu döner; hiç veri yoksa / bağımlılık yoksa None."""
    try:
        import pandas as pd
        import numpy as np
        from matplotlib.backends.backend_agg import FigureCanvasAgg
    except Exception as e:
        print(f"[EXCEL] Ping ozet atlandi (bagimlilik yok): {e}")
        return None

    valid = [p for p in (txt_paths or []) if p and os.path.exists(p)]
    if not valid:
        return None

    from common.runners.base import grk_style_filename
    fname = grk_style_filename("pingOzet", brand, model, firmware, client=node_name, ext="xlsx")
    out_path = os.path.join(out_dir or os.path.dirname(valid[0]), fname)
    start_ts = test_start_time or datetime.now()
    if isinstance(start_ts, str):
        try:
            start_ts = datetime.fromisoformat(start_ts)
        except ValueError:
            start_ts = datetime.now()

    try:
        with pd.ExcelWriter(out_path, engine="xlsxwriter") as writer:
            wb = writer.book
            header_fmt = wb.add_format({"bold": True, "bg_color": "#D3D3D3"})

            ws_stats = wb.add_worksheet("Statistics")
            for col_idx, key in enumerate(_PING_STATS_KEYS):
                ws_stats.write(0, col_idx, key, header_fmt)
                ws_stats.set_column(col_idx, col_idx, 18)

            ws_graphs = wb.add_worksheet("Graphs")
            graph_row = 1
            row = 1
            all_dfs = []

            for path in valid:
                try:
                    df, target_ip = _parse_ping_log(path)
                except Exception as e:
                    print(f"[EXCEL] Ping parse hatasi ({os.path.basename(path)}): {e}")
                    continue
                if df is None or df.empty:
                    continue

                stats = _ping_statistics(df)
                df = _ping_identify_anomalies(df)

                stats["source_computer"] = node_name
                stats["source_file"] = os.path.basename(path)
                stats["ip_version"] = _ip_version(stats.get("target_ip") or "")
                stats["test_start_time"] = start_ts.strftime("%Y-%m-%d %H:%M:%S")
                stats["test_end_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                for col_idx, key in enumerate(_PING_STATS_KEYS):
                    val = stats.get(key, "N/A")
                    ws_stats.write(row, col_idx, round(val, 2) if isinstance(val, float) else val)
                row += 1

                df["source_computer"] = node_name
                df["source_file"] = os.path.basename(path)
                all_dfs.append(df)

                # Grafik başarısız olursa Excel'i iptal etme — istatistik/data yine gitsin
                try:
                    fig = _ping_time_graph(df)
                    FigureCanvasAgg(fig)
                    img = BytesIO()
                    fig.savefig(img, format="png")
                    img.seek(0)
                    ws_graphs.write(graph_row, 0,
                                    f"{node_name} — Hedef IP: {stats.get('target_ip')}", header_fmt)
                    ws_graphs.insert_image(graph_row + 1, 0, "", {"image_data": img})
                    graph_row += 45
                except Exception as e:
                    print(f"[EXCEL] Ping grafik hatasi ({os.path.basename(path)}): {e}")

            if not all_dfs:
                raise ValueError("hicbir ping logunda veri yok")

            combined = pd.concat(all_dfs, ignore_index=True).replace([np.inf, -np.inf], np.nan)
            if len(combined) > _PING_MAX_DATA_ROWS:
                step = max(1, len(combined) // _PING_MAX_DATA_ROWS)
                view = combined.iloc[::step].reset_index(drop=True)
                print(f"[EXCEL] Ping Data downsample: {len(combined)} -> {len(view)} (step={step})")
            else:
                view = combined
            view.to_excel(writer, sheet_name="Data", index=False, na_rep="")
            ws_data = writer.sheets["Data"]
            for idx, col in enumerate(view.columns):
                width = max(len(str(col)),
                            view[col].astype(str).str.len().max() if not view.empty else 0)
                ws_data.set_column(idx, idx, min(width + 2, 100))

        print(f"[EXCEL] Ping ozet Excel uretildi: {os.path.basename(out_path)}")
        return out_path
    except Exception as e:
        print(f"[EXCEL] Ping ozet uretilemedi ({node_name}): {e}")
        # Yarım kalan dosya FTP/Telegram'a gitmesin
        try:
            os.remove(out_path)
        except OSError:
            pass
        return None


# ─────────────────────────────────────────────────────────────────────────────
# IPERF — iperf3 client log'undan grafik Excel'i (GRK karşılığı yok, FULL Servis'e
# özgü). Yalnızca [SUM] satırları kullanılır.
# ─────────────────────────────────────────────────────────────────────────────

def _iperf_bits_to_mbps(value: float, unit: str) -> float:
    """iperf3'ün 'Kbits/sec' / 'Mbits/sec' / 'Gbits/sec' / 'Tbits/sec' değerini
    Mbit/sn'ye normalize eder."""
    u = (unit or "").upper()
    if u.startswith("G"):
        return round(value * 1_000, 3)
    if u.startswith("T"):
        return round(value * 1_000_000, 3)
    if u.startswith("K"):
        return round(value / 1_000, 3)
    return round(value, 3)  # zaten Mbits


_IPERF_ROW_RE = re.compile(
    r"^\[\s*(SUM|\d+)\]\s+([\d.]+)\s*-\s*([\d.]+)\s+sec\s+"
    r"[\d.]+\s+[KMGT]?Bytes\s+"
    r"([\d.]+)\s+([KMGT]?bits/sec)"
    r"(?:\s+\d+)?"
    r"(?:\s+(sender|receiver))?",
    re.IGNORECASE,
)


def _parse_iperf_dt(raw: str):
    """'2026-06-15 14:42:21.475731' (Python str(datetime) çıktısı) parse eder."""
    raw = (raw or "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def _parse_iperf_log(file_path):
    """Bir iperf3 client logunu ayrıştırır: Komut satırından server/port/duration/
    parallel, Baslangic/Bitis satırlarından zaman damgası, [ID]/[SUM] satırlarından
    saniye başı throughput (Mbps) serisi.

    -P 1'de iperf3 [SUM] satırı basmaz; bu durumda tek akışın kendi satırları
    kullanılır."""
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        lines = f.readlines()
    text = "".join(lines)

    node_m = re.search(r"Node:\s*(\S+)", text)
    cmd_m = re.search(r"Komut:\s*(.+)", text)
    cmd = cmd_m.group(1).strip() if cmd_m else ""

    def _cmd_int(flag):
        m = re.search(rf"{flag}\s+(\d+)", cmd)
        return int(m.group(1)) if m else None

    server_m = re.search(r"-c\s+(\S+)", cmd)
    # Server logunda "-c" hiç yazmaz (yalnizca client'a ozgu parametre); bu
    # durumda sunucu IP'sini "local X port Y connected to" satirindan (kendi
    # dinledigi adres) turetiyoruz.
    local_m = re.search(r"local\s+([\d.]+)\s+port\s+\d+\s+connected to", text)
    server_ip = server_m.group(1) if server_m else (local_m.group(1) if local_m else None)
    start_m = re.search(r"Baslangic:\s*(.+)", text)
    end_m = re.search(r"Bitis:\s*(.+)", text)

    periodic, totals = [], []
    for line in lines:
        m = _IPERF_ROW_RE.match(line.strip())
        if not m:
            continue
        stream_id, t0, t1, rate_val, rate_unit, role = m.groups()
        row = {
            "stream_id": stream_id,
            "interval_start_sec": float(t0),
            "interval_end_sec": float(t1),
            "bitrate_mbps": _iperf_bits_to_mbps(float(rate_val), rate_unit),
            "role": role.lower() if role else None,
        }
        (totals if role else periodic).append(row)

    sum_periodic = [r for r in periodic if r["stream_id"] == "SUM"]
    series = sorted(sum_periodic if sum_periodic else periodic,
                    key=lambda r: r["interval_start_sec"])

    sum_totals = [r for r in totals if r["stream_id"] == "SUM"]
    final_totals = sum_totals if sum_totals else totals
    sender_mbps = next((r["bitrate_mbps"] for r in final_totals if r["role"] == "sender"), None)
    receiver_mbps = next((r["bitrate_mbps"] for r in final_totals if r["role"] == "receiver"), None)

    # Server logunda "-t" hiç yazmaz (client'a özgü parametre); bu durumda
    # süreyi periyodik satırların kapsadığı gerçek aralıktan türetiyoruz.
    requested_duration = _cmd_int("-t")
    measured_duration = round(max((r["interval_end_sec"] for r in series), default=0), 2) if series else None

    return {
        "node_name": node_m.group(1) if node_m else None,
        "server_ip": server_m.group(1) if server_m else None,
        "port": _cmd_int("-p"),
        "duration": requested_duration,
        "measured_duration_sec": measured_duration,
        "parallel": _cmd_int("-P") or 1,
        "sender_mbps": sender_mbps,
        "receiver_mbps": receiver_mbps,
        "has_final_summary": bool(final_totals),
        "test_start_time": _parse_iperf_dt(start_m.group(1)) if start_m else None,
        "test_end_time": _parse_iperf_dt(end_m.group(1)) if end_m else None,
        "series": series,
    }


def iperf_summary_excel(txt_paths: list[str], node_name: str, server_node_name=None,
                        brand=None, model=None, firmware=None,
                        out_dir: str | None = None) -> str | None:
    """Bir düğümün iperf3 client log(lar)ını grafik Excel'e çevirir — her log için
    bir "Grafik" (saniye başı [SUM] throughput çizgi grafiği) ve bir "DataLog"
    (ham Mbps değerleri + ortalama/sender/receiver özeti) sayfası.

    İsimlendirme: fullServis_iperfOzet_<bilgisayar>_<marka>_<model>_<fw>_<zaman>.xlsx

    Üretilen Excel'in yolunu döner; hiç geçerli log yoksa / bağımlılık yoksa None."""
    try:
        import xlsxwriter
    except Exception as e:
        print(f"[EXCEL] Iperf ozet atlandi (bagimlilik yok): {e}")
        return None

    valid = [p for p in (txt_paths or []) if p and os.path.exists(p)]
    if not valid:
        return None

    from common.runners.base import grk_style_filename
    fname = grk_style_filename("iperfOzet", brand, model, firmware, client=node_name, ext="xlsx")
    out_path = os.path.join(out_dir or os.path.dirname(valid[0]), fname)

    multi = len(valid) > 1
    wb = xlsxwriter.Workbook(out_path)
    produced = 0

    for idx, path in enumerate(valid, start=1):
        try:
            parsed = _parse_iperf_log(path)
        except Exception as e:
            print(f"[EXCEL] Iperf parse hatasi ({os.path.basename(path)}): {e}")
            continue

        suffix = f" {idx}" if multi else ""
        data_name = f"DataLog{suffix}"
        ws_data = wb.add_worksheet(data_name)
        ws_data.set_column(2, 2, 16)
        ws_data.set_column(3, 3, 22)
        values = [r["bitrate_mbps"] for r in parsed["series"]]
        for i, v in enumerate(values):
            ws_data.write(i, 0, v)

        def _cell(val):
            return val if val is not None else "N/A"

        def _rate_cell(val):
            """sender/receiver icin: deger yoksa, final ozet blogu VARDI ama bu
            rol icin satir hic yazilmamissa (test kesintiye ugramis olabilir)
            bunu ayirt eden bir not doner — duz "N/A" degil."""
            if val is not None:
                return val
            if parsed.get("has_final_summary"):
                return "N/A (final satir yazilmadi — test kesintiye ugramis olabilir)"
            return "N/A"

        # Sure: once client'in istedigi (-t) deger; o log'da yoksa (server
        # logunda -t hic gecmez) periyodik satirlardan olculen gercek sureye
        # duser — boylece Sure hicbir zaman bos kalmaz (veri varsa).
        duration_cell = parsed.get("duration")
        if duration_cell is None:
            duration_cell = parsed.get("measured_duration_sec")

        # DB'deki iperf_test tablosuyla eşleşen alanlar (session_id/test_name/
        # ftp_file_path/created_at hariç — onlar log'dan değil, DB yazma anında
        # oturum/FTP bağlamından gelir).
        info_rows = [
            ("Node", _cell(parsed.get("node_name") or node_name)),
            ("Server Node", _cell(server_node_name)),
            ("Server IP", _cell(parsed.get("server_ip"))),
            ("Port", _cell(parsed.get("port"))),
            ("Parallel", _cell(parsed.get("parallel"))),
            ("Sure (sn)", _cell(duration_cell)),
            ("Baslangic", _cell(parsed["test_start_time"].strftime("%Y-%m-%d %H:%M:%S")
                                if parsed.get("test_start_time") else None)),
            ("Bitis", _cell(parsed["test_end_time"].strftime("%Y-%m-%d %H:%M:%S")
                            if parsed.get("test_end_time") else None)),
            ("Sender (Mbps)", _rate_cell(parsed.get("sender_mbps"))),
            ("Receiver (Mbps)", _rate_cell(parsed.get("receiver_mbps"))),
        ]
        if values:
            info_rows.append(("Ortalama (Mbps)", round(sum(values) / len(values), 2)))
        for r, (label, val) in enumerate(info_rows):
            ws_data.write(r, 2, label)
            ws_data.write(r, 3, val)
        note_row = len(info_rows)

        if values:
            requested = parsed.get("duration")
            if requested and abs(len(values) - requested) > 2:
                msg = f"UYARI: test {requested}sn surmesi gerekirken {len(values)}sn olarak algilandi"
                print(f"[EXCEL] Iperf {os.path.basename(path)}: {msg}")
                ws_data.write(note_row, 2, msg)
            ws_graph = wb.add_worksheet(f"Grafik{suffix}")
            chart = wb.add_chart({"type": "line"})
            chart.add_series({
                "values": [data_name, 0, 0, len(values) - 1, 0],
                "line": {"color": "red"},
            })
            title = f"{parsed.get('node_name') or node_name} -> {parsed.get('server_ip')}:{parsed.get('port')}"
            chart.set_title({"name": f"iperf3 Throughput — {title}"})
            chart.set_x_axis({"name": "Saniye"})
            chart.set_y_axis({"name": "Mbps"})
            chart.set_style(5)
            chart.set_size({"x_scale": 2, "y_scale": 1.5})
            ws_graph.insert_chart("A1", chart)
        else:
            note = "Veri yok — baglanti kurulamamis olabilir"
            ws_data.write(note_row, 2, note)
            print(f"[EXCEL] Iperf {os.path.basename(path)}: {note}")

        produced += 1

    if not produced:
        wb.close()
        try:
            os.remove(out_path)
        except OSError:
            pass
        return None

    wb.close()
    print(f"[EXCEL] Iperf ozet Excel uretildi: {os.path.basename(out_path)}")
    return out_path


# ─────────────────────────────────────────────────────────────────────────────
# WIFI — GRK parse.py (create_master_excel) + chart.py (create_combined_chart) portu
# ─────────────────────────────────────────────────────────────────────────────

def _wifi_safe_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _wifi_safe_int(val):
    try:
        return int(val)
    except (TypeError, ValueError):
        return None


def _wifi_parse_time(time_str):
    """'Fri Apr 17 17:35:44 2026' gibi ctime formatını datetime'a çevirir."""
    if not time_str:
        return None
    s = time_str.strip()
    if not s or s.lower() == "time":
        return None
    for fmt in ("%a %b %d %H:%M:%S %Y", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _wifi_mode(lst):
    return Counter(lst).most_common(1)[0][0] if lst else None


def _wifi_avg(lst):
    nums = [x for x in lst if x is not None]
    return round(sum(nums) / len(nums), 2) if nums else None


def _wifi_create_master_excel(log_data, output_file):
    """GRK create_master_excel portu — 'Veriler' + 'Özet' sayfalarını yazar,
    özet istatistik sözlüğünü döner."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    total_sheet = wb.create_sheet("Veriler")

    signal_vals, rx_vals, tx_vals, cpu_vals, ram_vals = [], [], [], [], []
    channel_vals, protocol_vals, state_vals, time_vals = [], [], [], []

    for line in log_data[1:]:
        parts = line.split("|")
        if len(parts) < 3:
            continue

        time_str = parts[0].strip().replace("-----", "").strip()
        if time_str.lower() == "time" or "wlan_state" in parts[1].lower():
            continue
        values = parts[1].strip().split(", ")
        cpuram = parts[2].strip().split(", ")

        power_state = cpuram[2].strip() if len(cpuram) > 2 else ""
        battery = cpuram[3].split("%")[0].strip() if len(cpuram) > 3 else ""

        conn_str = values[1].lower() if len(values) > 1 else ""
        if ("connected" in conn_str or "bağlı değil" in conn_str
                or "kesildi" in conn_str or "disconnected" in conn_str):
            if ("disconnected" in conn_str or "bağlı değil" in conn_str
                    or "değil" in conn_str or "kesildi" in conn_str):
                state_status = 0
            else:
                state_status = 1
        elif "bağlı" in conn_str:
            state_status = 1
        else:
            state_status = 2

        signal_raw = values[2].split("%")[0].strip() if len(values) > 2 else ""
        signal_status = signal_raw
        for i in range(0, 101):
            if signal_raw == f"{i}":
                signal_status = i
                break

        rx_raw = values[3].strip() if len(values) > 3 else ""
        rx_status = rx_raw
        for i in range(1, 1001):
            if rx_raw == f"{i}":
                rx_status = i
                break

        tx_raw = values[4].strip() if len(values) > 4 else ""
        tx_status = tx_raw
        for i in range(1, 1001):
            if tx_raw == f"{i}":
                tx_status = i
                break

        ch_raw = values[5].strip() if len(values) > 5 else ""
        ch_status = ch_raw
        for i in range(1, 25):
            if ch_raw == f"{i}":
                ch_status = i
                break

        protocol = values[6].strip() if len(values) > 6 else ""
        protocol_map = {"802.11be": 6, "802.11ax": 5, "802.11ac": 4,
                        "802.11n": 3, "802.11g": 2, "802.11b": 1}
        protocol_status = protocol_map.get(protocol, 0)

        cpu_raw = cpuram[0].split("%")[0].strip() if len(cpuram) > 0 else ""
        cpu_status = cpu_raw
        for i in range(0, 101):
            v = i / 10
            if cpu_raw == f"{v}":
                cpu_status = str(v)
                break

        ram_raw = cpuram[1].split("%")[0].strip() if len(cpuram) > 1 else ""
        ram_status = ram_raw
        for i in range(0, 101):
            v = i / 10
            if ram_raw == f"{v}":
                ram_status = str(v)
                break

        total_sheet.append([
            time_str, conn_str, state_status,
            signal_status, rx_status, tx_status,
            ch_status, protocol, protocol_status,
            cpu_status, ram_status, power_state, battery,
        ])

        parsed_t = _wifi_parse_time(time_str)
        if parsed_t:
            time_vals.append(parsed_t)
        state_vals.append(state_status)
        sig = _wifi_safe_int(signal_status) if isinstance(signal_status, int) else _wifi_safe_float(signal_status)
        signal_vals.append(sig)
        rx_vals.append(float(rx_status) if isinstance(rx_status, int) else _wifi_safe_float(rx_status))
        tx_vals.append(float(tx_status) if isinstance(tx_status, int) else _wifi_safe_float(tx_status))
        cpu_vals.append(_wifi_safe_float(cpu_status))
        ram_vals.append(_wifi_safe_float(ram_status))
        if ch_status:
            channel_vals.append(str(ch_status))
        if protocol:
            protocol_vals.append(protocol)

    # Sütun genişlikleri + başlık etiketleri (GRK ile birebir)
    widths = {"A": 27, "B": 10, "C": 11, "D": 15, "E": 12, "F": 12,
              "G": 10, "H": 10, "I": 13, "J": 15, "K": 15, "L": 15, "M": 15}
    for col, w in widths.items():
        total_sheet.column_dimensions[col].width = w

    bold = Font(bold=True)
    total_sheet["C1"] = "State Value\n   1 : connected\n   0 : disconnected\n   2 : other"
    total_sheet["D1"] = "Signal Level (%)"
    total_sheet["J1"] = "CPU Usage (%)"
    total_sheet["K1"] = "RAM Usage (%)"
    total_sheet["I1"] = "Protocol Value\n   5 : 802.11ax\n    4 : 802.11ac\n    3 : 802.11n\n    2 : 802.11g\n    1 : 802.11b"
    total_sheet["L1"] = "Power State"
    total_sheet["M1"] = "Battery (%)"
    for col in "ABCDEFGHIJKLM":
        total_sheet[f"{col}1"].font = bold
    for col in total_sheet.iter_cols(min_col=1, max_col=12, min_row=2):
        for cell in col:
            cell.alignment = Alignment(horizontal="center")

    # Özet istatistikler
    total_samples = len(state_vals)
    connected = state_vals.count(1)
    disconnected = state_vals.count(0)
    dominant_channel = _wifi_mode(channel_vals)
    dominant_protocol = _wifi_mode(protocol_vals)
    sig_nums = [x for x in signal_vals if x is not None]
    avg_signal = round(sum(sig_nums) / len(sig_nums), 2) if sig_nums else None
    min_signal = min(sig_nums) if sig_nums else None
    max_signal = max(sig_nums) if sig_nums else None
    avg_rx, avg_tx = _wifi_avg(rx_vals), _wifi_avg(tx_vals)
    avg_cpu, avg_ram = _wifi_avg(cpu_vals), _wifi_avg(ram_vals)
    test_start_dt = time_vals[0] if time_vals else None
    test_end_dt = time_vals[-1] if time_vals else None

    summary_stats = {
        "total_samples": total_samples,
        "connected_count": connected,
        "disconnected_count": disconnected,
        "channel": dominant_channel,
        "wifi_protocol": dominant_protocol,
        "avg_signal_percentage": avg_signal,
        "min_signal_percentage": float(min_signal) if min_signal is not None else None,
        "max_signal_percentage": float(max_signal) if max_signal is not None else None,
        "avg_rx_rate": avg_rx,
        "avg_tx_rate": avg_tx,
        "avg_cpu_usage": avg_cpu,
        "avg_ram_usage": avg_ram,
        "test_start_time": test_start_dt,
        "test_end_time": test_end_dt,
    }

    # Özet sayfası
    ozet = wb.create_sheet("Özet")
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    rows = [
        ("Parametre", "Değer"),
        ("Toplam Örnek (sn)", total_samples),
        ("Bağlı Süre (sn)", connected),
        ("Kopan Süre (sn)", disconnected),
        ("Kanal", dominant_channel),
        ("Wi-Fi Protokolü", dominant_protocol),
        ("Ort. Sinyal (%)", avg_signal),
        ("Min. Sinyal (%)", min_signal),
        ("Max. Sinyal (%)", max_signal),
        ("Ort. RX Hız (Mbps)", avg_rx),
        ("Ort. TX Hız (Mbps)", avg_tx),
        ("Ort. CPU (%)", avg_cpu),
        ("Ort. RAM (%)", avg_ram),
        ("Test Başlangıç", test_start_dt.strftime("%Y-%m-%d %H:%M:%S") if test_start_dt else None),
        ("Test Bitiş", test_end_dt.strftime("%Y-%m-%d %H:%M:%S") if test_end_dt else None),
    ]
    for r_idx, (param, val) in enumerate(rows, start=1):
        a = ozet.cell(row=r_idx, column=1, value=param)
        b = ozet.cell(row=r_idx, column=2, value=val)
        a.font = bold
        if r_idx == 1:
            a.fill = header_fill
            b.fill = header_fill
        a.alignment = Alignment(horizontal="left")
        b.alignment = Alignment(horizontal="center")
    ozet.column_dimensions["A"].width = 25
    ozet.column_dimensions["B"].width = 20

    wb.remove(wb.active)   # otomatik açılan boş "Sheet"i kaldır
    wb.save(output_file)
    return summary_stats


def _wifi_create_combined_chart(master_file):
    """GRK chart.py create_combined_chart portu — 'Veriler'den 10 alt grafik üretip
    'Grafikler' sayfası olarak Excel'e gömer."""
    import matplotlib
    matplotlib.use("Agg")
    from matplotlib.figure import Figure
    from matplotlib.lines import Line2D
    import matplotlib.dates as mdates
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage

    df = pd.read_excel(master_file, sheet_name="Veriler")
    col_name = df.columns[0]
    df[col_name] = pd.to_datetime(df[col_name], errors="coerce")
    df = df.dropna(subset=[col_name])
    if df.empty:
        return

    show_markers = len(df) <= 500

    metrics = [
        ("State Value",      df.columns[2]),
        ("Signal Level (%)", df.columns[3]),
        ("Transmit Rate",    df.columns[5]),
        ("Receive Rate",     df.columns[4]),
        ("Wi-Fi Channel",    df.columns[6]),
        ("Wi-Fi Protocol",   df.columns[8]),
        ("CPU Usage (%)",    df.columns[9]),
        ("RAM Usage (%)",    df.columns[10]),
        ("Power State",      df.columns[11]),
        ("Battery (%)",      df.columns[12]),
    ]

    fig = Figure(figsize=(16, 14))
    axes = fig.subplots(5, 2, sharex=True).flatten()

    for ax, (label, col) in zip(axes, metrics):
        if col in df:
            kwargs = dict(linewidth=1, rasterized=True)
            if show_markers:
                kwargs.update(marker="o", markersize=3)
            try:
                ax.plot(df.iloc[:, 0], df[col].astype(float), **kwargs)
            except Exception:
                pass
        ax.set_title(label)
        ax.grid(False)

        if label == "State Value":
            handles = [Line2D([0], [0], color="black", marker="o", linestyle="None", label=t)
                       for t in ("1 – connected", "0 – disconnected", "2 – different")]
            ax.legend(handles=handles, title="State Mapping", loc="best", fontsize="small")
        elif label == "Wi-Fi Protocol":
            handles = [Line2D([0], [0], color="black", marker="o", linestyle="None", label=t)
                       for t in ("6 – 802.11be", "5 – 802.11ax", "4 – 802.11ac",
                                 "3 – 802.11n", "2 – 802.11g", "1 – 802.11b", "0 – other")]
            ax.legend(handles=handles, title="Protocol Mapping", loc="upper right",
                      bbox_to_anchor=(1, 1), fontsize="x-small", markerscale=0.8, frameon=False)
        elif label == "Power State":
            handles = [Line2D([0], [0], color="black", marker="o", linestyle="None", label=t)
                       for t in ("1 – true", "0 – false")]
            ax.legend(handles=handles, title="Power State", loc="best", fontsize="small")
        else:
            ax.legend([label], loc="best", fontsize="small")

    for ax in axes[-2:]:
        ax.set_xlabel("Zaman")
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M:%S"))
        ax.xaxis.set_major_locator(mdates.AutoDateLocator(maxticks=8))
    fig.autofmt_xdate(rotation=30, ha="right")
    fig.tight_layout()

    img = BytesIO()
    fig.savefig(img, format="png", dpi=150)
    img.seek(0)

    wb = load_workbook(master_file)
    if "Grafikler" in wb.sheetnames:
        del wb["Grafikler"]
    ws = wb.create_sheet("Grafikler")
    xlimg = XLImage(img)
    xlimg.anchor = "B2"
    ws.add_image(xlimg)
    wb.save(master_file)


def wifi_log_to_excel(txt_path: str) -> str | None:
    """Wifi log .txt → GRK formatında .xlsx (Veriler + Özet + Grafikler).
    Üretilen Excel yolunu döner; veri yoksa/araç yoksa None."""
    try:
        import openpyxl  # noqa: F401
        import pandas  # noqa: F401
        import matplotlib  # noqa: F401
    except Exception as e:
        print(f"[EXCEL] Wifi Excel atlandi (bagimlilik yok): {e}")
        return None

    try:
        with open(txt_path, "r", encoding="utf-8", errors="replace") as f:
            log_data = [ln.strip() for ln in f.readlines()]
    except Exception as e:
        print(f"[EXCEL] Wifi log okunamadi: {e}")
        return None

    out_path = os.path.splitext(txt_path)[0] + ".xlsx"
    try:
        summary = _wifi_create_master_excel(log_data, out_path)
        if not summary or not summary.get("total_samples"):
            print(f"[EXCEL] Wifi logunda veri yok, Excel uretilmedi: {os.path.basename(txt_path)}")
            # Boş/yetersiz: yarım Excel kalmasın
            try:
                os.remove(out_path)
            except OSError:
                pass
            return None
        _wifi_create_combined_chart(out_path)
        print(f"[EXCEL] Wifi Excel uretildi: {os.path.basename(out_path)}")
        return out_path
    except Exception as e:
        print(f"[EXCEL] Wifi Excel uretim hatasi: {e}")
        return None
